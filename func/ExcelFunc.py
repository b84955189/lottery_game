#!/usr/bin/env python3
# -*- encoding: utf-8 -*-
"""
@File    :   ExcelFunc.py    
@Contact :   lking@lking.icu
@Author :    Jason
@Date :      2022/4/28 19:36
@Description  Python version-3.10
"""

from openpyxl import load_workbook

from func import CommonTools


# 表 标题 数据行位置 固定为 1
HEADER_ROW = 1


def get_workbook(file_path):
    return load_workbook(file_path)


def close_workbook(workbook):
    """
    关闭Excel工作簿对象
    @param workbook: Excel工作簿对象
    @return:
    """
    if workbook is not None:
        workbook.close()


def get_data_from_sheet(sheet):
    """
    获取 学生 数据
    openpyxl 的 行 与 列 都是从 1 开始
    @param sheet: 表对象
    @return: 姓名数据 - dict
    """
    names = {}
    # 读表头->找‘姓名’表头 的列
    row_column = 1
    while "姓名" != sheet.cell(HEADER_ROW, row_column).value:
        # 防止死循环
        if row_column == 999:
            raise Exception('数据文件中没有姓名数据！')
        row_column += 1
    # 读‘姓名’数据
    # 跳过表头
    num = 1
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, row_column).value
        # 判断错误数据
        if CommonTools.is_empty_or_none(name):
            continue
        # 为每个学生添加编号
        names[name] = num
        num += 1

    return names

